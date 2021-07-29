from openpyxl import workbook, load_workbook
from openpyxl.utils import get_column_letter
from time import strftime
import codecs

wb = load_workbook('horario.xlsx')
ws = wb.active

tabela = {'Mon':1,'Tue':4,'Wed':7,'Thu':10,'Fri':13}


with codecs.open('dados.txt', encoding='utf-8', mode='r') as ler:
    Dados = ler.read().splitlines()

Camil = Dados[5]

colunaHora = tabela.get(strftime('%a'))
if Camil: colunaAula = colunaHora+1
else: colunaAula = colunaHora+2


def compara_hora(CH,CA):
    hora = 650

    for row in range(2,13):
        horario = ws[get_column_letter(CH)+str(row)].value

        if ws[get_column_letter(CA)+str(row)].value == None: 
            return ws[get_column_letter(CA)+str(row-1)].value

        elif hora > horario: pass
        else: return ws[get_column_letter(CA)+str(row)].value


aulaAtual = compara_hora(colunaHora,colunaAula).encode('utf-8')


with  open("dados.txt", "r") as arquivo:
    linhas = arquivo.readlines()

if len(linhas) == 7:
    del linhas[6]

with open("dados.txt", "w+") as arquivoNovo:
    for linha in linhas:
        arquivoNovo.write(linha)

with open('dados.txt', 'ab') as nota:
    nota.write(aulaAtual)

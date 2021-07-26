from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from time import strftime, sleep
import codecs

#Lê o horário e atualiza [6] de dados.txt com o nome da aula atual

wb = load_workbook('horario.xlsx')
ws = wb.active

tabela = {'Mon':1,'Tue':4,'Wed':7,'Thu':10,'Fri':13}

hora = int(strftime('%H%M'))
dia = strftime('%a')

with codecs.open('dados.txt', encoding='utf-8', mode='r') as ler:
    Dados = ler.read().splitlines()

Camil = Dados[5]

diaHora = tabela.get(dia)
if Camil: diaAula = diaHora+1
else: diaAula = diaHora+2



def compara_hora(CH,CA,C,F):
    letra = get_column_letter(CH)
    aula = get_column_letter(CA)
    for row in range(C,F+1):
        valor = ws[letra+str(row)].value
        if hora <= valor:
            return(ws[aula+str(row-1)].value)
        else: pass


aulaAtual = compara_hora(diaHora,diaAula,2,11).encode('utf-8')


with  open("dados.txt", "r") as arquivo:
    linhas = arquivo.readlines()

if len(linhas) == 7:
    del linhas[6]

with open("dados.txt", "w+") as arquivoNovo:
    for linha in linhas:
        arquivoNovo.write(linha)

with open('dados.txt', 'ab') as nota:
    nota.write(aulaAtual)


#Usa o Selenium para abrir o formulário e preencher com os dados do usuário

from selenium import webdriver
import codecs
from tkinter import messagebox, Tk

browser = webdriver.Chrome('chromedriver')
browser.get('https://docs.google.com/forms/d/e/1FAIpQLSf8IpgwW4DcIsvrKMp-UhffOTP5MMiSCd5ttxVvK6pFXXbwCA/viewform')

sleep(1.5)
    
with codecs.open('dados.txt', encoding='utf-8', mode='r') as ler:
    Dados = ler.read().splitlines()

Email = Dados[0]
Nome = Dados[1]
Numero = Dados[2]
Turma = Dados[3]
Grupamento = Dados[4]
Aula = Dados[6]


idBotao = {'301':17, '302':20, '303':23, '304':26, '305':29, "A":36,"B":39}

idCaixa = {
"Geografia":3,
"História":4,
"Matemática":5,
"literatura":6,
"Biologia":7,
"Arte":8,
"Educação física":9,
"Português":10,
"Sociologia":11,
"Filosofia":12,
"Química":13,
"Física":14,
"Redação":15,
"Inglês":16,
"Espanhol":17}


inputEmail = browser.find_element_by_xpath('//*[@id="mG61Hd"]/div[2]/div/div[2]/div[1]/div/div[1]/div[2]/div[1]/div/div[1]/input')
inputEmail.send_keys(Email)


inputNome = browser.find_element_by_xpath('//*[@id="mG61Hd"]/div[2]/div/div[2]/div[2]/div/div/div[2]/div/div[1]/div/div[1]/input')
inputNome.send_keys(Nome)


inputNumero = browser.find_element_by_xpath('//*[@id="mG61Hd"]/div[2]/div/div[2]/div[3]/div/div/div[2]/div/div[1]/div/div[1]/input')
inputNumero.send_keys(Numero) 


inputTurma = browser.find_element_by_xpath(f'//*[@id="i{idBotao.get(Turma)}"]/div[3]/div')
inputTurma.click()


inputGrupamento = browser.find_element_by_xpath(f'//*[@id="i{idBotao.get(Grupamento)}"]/div[3]/div')
inputGrupamento.click()


inputAula = browser.find_element_by_xpath('//*[@id="mG61Hd"]/div[2]/div/div[2]/div[6]/div/div/div[2]/div/div[1]/div[1]/div[1]')
inputAula.click()


sleep(0.5)
try:
    inputAula = browser.find_element_by_xpath(f'//*[@id="mG61Hd"]/div[2]/div/div[2]/div[6]/div/div/div[2]/div/div[2]/div[{idCaixa.get(Aula)}]/span')
except: 
    Tk().withdraw()
    messagebox.showerror('Erro', 'Nenhuma aula encontrada')
    browser.quit()
    exit()

    
inputAula.click()

sleep(2)

enviar = browser.find_element_by_xpath('//*[@id="mG61Hd"]/div[2]/div/div[3]/div[3]/div/div/span')
enviar.click()